import openpyxl
import json

wb_val = openpyxl.load_workbook('Book1.xlsx', data_only=True)
ws_val = wb_val['Fan Lathe']

days_names = ['Fri', 'Sat', 'Sun', 'Mon', 'Tue', 'Wed', 'Thu']

clean_rows = []

for r in range(6, 99):
    row_idx = r - 6
    day_num = (row_idx // 3) + 1 # 1 to 31
    is_day_first_row = (row_idx % 3) == 0
    
    date_str = f"{day_num:02d}-May-2026"
    day_name = days_names[(day_num - 1) % 7]
    
    row_dict = {'row': r}
    
    # Col A: Date
    row_dict['A'] = {
        'val': date_str if is_day_first_row else None,
        'formula': '=AM3' if r == 6 else (f'=A{r-3}+1' if is_day_first_row else None)
    }
    
    # Col B: Day
    row_dict['B'] = {
        'val': day_name if is_day_first_row else None,
        'formula': f'=A{r}' if is_day_first_row else None
    }
    
    # Col C: Shift
    shift_val = ws_val.cell(r, 3).value
    row_dict['C'] = {
        'val': 'Morning' if not shift_val or str(shift_val).startswith('=' ) else str(shift_val),
        'formula': None
    }
    
    # Col D: Machine Name
    machine_val = ws_val.cell(r, 4).value
    if not machine_val:
        machine_val = 'Rotor (CNC)' if row_idx % 3 == 0 else ('Bottom Cover (CNC)' if row_idx % 3 == 1 else 'Top Cover (CNC)')
    row_dict['D'] = {
        'val': str(machine_val),
        'formula': None
    }
    
    # Col E: Machine Capacity
    cap_val = ws_val.cell(r, 5).value
    if cap_val is None:
        cap_val = 2000 if row_idx % 3 == 0 else 3000
    row_dict['E'] = {
        'val': cap_val,
        'formula': None
    }
    
    # Col F: Actual Prd
    act_val = ws_val.cell(r, 6).value
    row_dict['F'] = {
        'val': act_val if act_val is not None else None,
        'formula': None
    }
    
    # Col G: Rejection
    rej_val = ws_val.cell(r, 7).value
    row_dict['G'] = {
        'val': rej_val if rej_val is not None else None,
        'formula': None
    }
    
    # Col H: Planned Prd Time
    plan_val = ws_val.cell(r, 8).value
    row_dict['H'] = {
        'val': plan_val if plan_val is not None else 660,
        'formula': None
    }
    
    # Col I: Expected DownTime
    row_dict['I'] = {
        'val': 30 if row_dict['E']['val'] else 0,
        'formula': f'=IF(E{r}<>"",30,0)'
    }
    
    # Col J: Total Prd Run Time
    row_dict['J'] = {
        'val': 660,
        'formula': f'=H{r}-AH{r}'
    }
    
    # Cols K to AG (23 Downtimes)
    for c in range(11, 34):
        col_letter = openpyxl.utils.get_column_letter(c)
        dt_val = ws_val.cell(r, c).value
        row_dict[col_letter] = {
            'val': dt_val if dt_val is not None and dt_val != 0 else None,
            'formula': None
        }
        
    # Col AH: Total Down Time
    row_dict['AH'] = {
        'val': 0,
        'formula': f'=SUM(K{r}:AG{r})'
    }
    
    # Col AI: Availability (%)
    row_dict['AI'] = {
        'val': 1.0,
        'formula': f'=IFERROR(J{r}/H{r},"0")'
    }
    
    # Col AJ: Performance (%)
    row_dict['AJ'] = {
        'val': 0.0,
        'formula': f'=IFERROR(F{r}/E{r},"0")'
    }
    
    # Col AK: Quality (%)
    row_dict['AK'] = {
        'val': 1.0,
        'formula': f'=IFERROR(F{r}/(F{r}+G{r}),"0")'
    }
    
    # Col AL: OEE (%)
    row_dict['AL'] = {
        'val': 0.0,
        'formula': f'=IFERROR(AK{r}*AJ{r}*AI{r},"0")'
    }
    
    # Col AM: Remarks
    rem_val = ws_val.cell(r, 39).value
    row_dict['AM'] = {
        'val': rem_val if rem_val is not None else '',
        'formula': None
    }
    
    clean_rows.append(row_dict)

with open('initial_sheet_data.js', 'w', encoding='utf-8') as f:
    f.write('const INITIAL_EXCEL_ROWS = ' + json.dumps(clean_rows, ensure_ascii=False, indent=2) + ';')

print(f"Successfully wrote {len(clean_rows)} rows to initial_sheet_data.js!")
